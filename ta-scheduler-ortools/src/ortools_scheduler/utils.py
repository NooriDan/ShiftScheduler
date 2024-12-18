from typing import Dict
from datetime import datetime
import csv
import openpyxl
from pathlib import Path
import pandas as pd

# Custom Imports 
from .domain import TA, Shift, Schedule
from .solver import Scheduler


class DataConstructor:
    def __init__(self, ta_csv_path, shift_csv_path, availability_folder, load=True):
        self.ta_csv_path = ta_csv_path
        self.shift_csv_path = shift_csv_path
        self.availability_folder = availability_folder
        # Data holders
        self.ta_data    = None  # Placeholder for TA data    (pandas frame)
        self.shift_data = None  # Placeholder for Shift data (pandas frame)
        self.schedule   = None
        self.avialability_matrix: Dict[str, Dict[int, int]] = {}
        if load:
            self.create()

    def create(self):
        # Construction routine
        self.schedule = Schedule()
        self.load_data()
        self.create_ta_objects()
        self.create_shift_objects()
        self.extract_availabilities()
        if (self.validate_ta_availability()):
            self.update_ta_availability()
            print(f"Data succussfully extracted :)")
        else: 
            print("\n========================================")
        return self.schedule
    
    def extract_availabilities(self):
        print(self.availability_folder)
        filenames = self._find_xlsx_files(self.availability_folder)

        for file in filenames:
            mac_id, availability = self._extract_availability_from_xlsx(file)
            self.avialability_matrix[mac_id] = availability
        
        return self.avialability_matrix
    
    def validate_ta_availability(self, log_error=True):
        missing_ta = 0
        for ta in self.schedule.tas:
            if self.avialability_matrix.get(ta.macid) is None:
                missing_ta += 1
                if log_error:
                    print("========================================")
                    print(f"Found no record for TA: {ta.name}")
                    print(f"MacID = {ta.macid}")
                    print(f"missing {missing_ta}/{len(self.schedule.tas)} forms so far")
                    print("========================================")
            else:
                count_available = sum(1 for availability in self.avialability_matrix[ta.macid] if availability != "Unavailable")
                if count_available < ta.req_shift_per_week:
                    if log_error:
                        print(f"TA {ta.id} has insufficient availability.")
                        print(f"Available shifts: {count_available}")
                        print(f"Required shifts: {ta.req_shift_per_week}")

                    return False
        return not missing_ta
    
    def update_ta_availability(self):
        mapping: Dict[str, int] = {
            "Unavailable" : -1,
            "Desired"     : 1,
            "Undesired"   : 0
        }
        if (self.validate_ta_availability(log_error=False)):
            for ta in self.schedule.tas:
                ta.availability_as_dict = self.avialability_matrix.get(ta.macid)
                for shift in self.schedule.shifts:
                    ta.availability_as_array_int.append(mapping.get(ta.availability_as_dict.get(shift.series)))
            return True
        return False
    
    def _extract_availability_from_xlsx(self, file_path: str, availability_tbl_name: str = "availability_tbl"):
        # Load the workbook and select the sheet
        wb = openpyxl.load_workbook(file_path)
        ws = wb['availability']
        
        # Extracting TA details from the top of the sheet
        mac_id = ws['B1'].value
        ta_name = ws['B2'].value
        ta_id   = ws['B3'].value

        # # Print table names in the sheet
        # print("Tables in the sheet:", ws.tables.keys())

        # Check if availability_tbl_name exists
        if availability_tbl_name not in ws.tables:
            print("Table 'availability_tbl' not found.")

        tbl = ws.tables[availability_tbl_name]
        
        # Extract the table range
        table_range = tbl.ref  # e.g., "A5:C10"
        # print(f"Table '{availability_tbl_name}' range: {table_range}")

        # Extract the rows from the table range
        rows = ws[table_range]

        # Convert rows into a list of lists
        data = [[cell.value for cell in row] for row in rows]

        # Convert the data into a pandas DataFrame
        headers = data[0]       # First row contains headers
        table_df = pd.DataFrame(data[1:], columns=headers)  # Remaining rows are data
        
        availability = {}
        for index, row in table_df.iterrows():
            shift = row['Shift Series']
            shift_availability = row['Availability']
            availability[shift] = shift_availability   

        # print(availability)
        return mac_id, availability

    def _find_csv_files(self, directory):
        path = Path(directory)
        return [str(file) for file in path.glob("*.csv") if not file.name.startswith('~$')]  # Exclude files starting with ~$ and no recursive search

    def _find_xlsx_files(self, directory):
        path = Path(directory)
        return [str(file) for file in path.glob("*.xlsx") if not file.name.startswith('~$')]  # Exclude files starting with ~$ and no recursive search

    def load_data(self):
        self.ta_data    = pd.read_csv(self.ta_csv_path)
        self.shift_data = pd.read_csv(self.shift_csv_path)

    def create_ta_objects(self):
        for index, row in self.ta_data.iterrows():
            ta = TA(
                id= index,
                macid=row['macid'],
                name=row['name'],
                req_shift_per_week=row['req_shift_per_week']
            )
            self.schedule.tas.append(ta)

    def create_shift_objects(self):
        for index, row in self.shift_data.iterrows():
            shift = Shift(
                id = index,
                name = row['name'],
                series = row['series'],
                day_of_week = row['day_of_week'],
                date = datetime.strptime(row['date'], '%Y-%m-%d').date(),
                start_time = datetime.strptime(row['start_time'], '%H:%M').time(),
                req_ta_per_shift = row['req_ta_per_shift']
            )
            self.schedule.shifts.append(shift)



class Schedule_Utils:
    def __init__(self, output_dir: str = "outputs"):
        self.output_dir = output_dir

    # Reporting Utils
    def group_shifts_by_day_of_week(self, schedule_solution: Scheduler):
        workday_shifts = {}
        for shift in schedule_solution.schedule.shifts:
            if shift.day_of_week not in workday_shifts:
                workday_shifts[shift.day_of_week] = []
            workday_shifts[shift.day_of_week].append(shift)

        # Sort shifts for each workday based on start_time
        for day in workday_shifts:
            workday_shifts[day] = sorted(workday_shifts[day], key=lambda s: s.start_time)
        return workday_shifts

    def report(self, schedule_solution: Scheduler, filename='ta_schedule_report.csv'):
        # Step 1: Group shifts by day_of_week and sort them by start_time
        filename = f"{self.output_dir}/{filename}"
        workday_shifts = self.group_shifts_by_day_of_week(schedule_solution)

        # Step 2: Prepare the header for the CSV file
        days_of_week = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        header = ['Shift/TA'] + days_of_week  # First column will be shift names, rest are days

        # Step 3: Determine the max number of rows across all workdays
        max_rows = max(len(shifts) for shifts in workday_shifts.values())  # Find the max number of shifts per day

        # Step 4: Open the CSV file for writing
        with open(filename, mode='w', newline='') as file:
            writer = csv.writer(file)

            # Step 5: Write the header row to the CSV
            writer.writerow(header)

            # Step 6: Write the shift rows
            for row_index in range(max_rows):
                row = ['']  # The first column is for shift names (e.g., Lab_0_0)
                for day in days_of_week:
                    shifts = workday_shifts.get(day, [])
                    if row_index < len(shifts):
                        shift = shifts[row_index]
                        # Collect assigned TAs for this shift
                        assigned_tas = [f"{ta.id}_{ta.name}" for ta in shift.assigned_tas]
                        # Prepare the cell content with shift name and assigned TAs
                        cell_content = f"{shift.series}_{shift.id}\n" + "\n".join(assigned_tas) if assigned_tas else f"{shift.series}_{shift.id}\nNo TAs assigned"
                    else:
                        cell_content = ""  # Empty cell if no shift for this row
                    row.append(cell_content)

                # Step 7: Write the row to the CSV
                writer.writerow(row)

        print(f"Report has been saved to {filename}")

    def report_timeseries(self, schedule_solution: Scheduler, filename='ta_schedule_report_timeseries.csv'):
        filename = f"{self.output_dir}/{filename}"
        # Step 1: Extract and sort the shifts based on start time
        shifts_sorted_by_start_time = sorted(schedule_solution.schedule.shifts, key=lambda s: s.start_time)

        # Step 2: Prepare the header for the CSV file using the start time of each shift
        header = ['TA Name/row']  # First column will be TA names and IDs
        for shift in shifts_sorted_by_start_time:
            # Format date and time as a string (assumes shift has a start_time field)
            date_combined  = datetime.combine(shift.date, shift.start_time)
            start_time_str = date_combined.strftime('%B %d %H:%M')
            to_append = f"{shift.series} ({start_time_str})"
            header.append(to_append)

        # Step 3: Determine the max number of rows (based on the most TAs assigned to a shift)
        max_rows = max(len(shift.assigned_tas) for shift in schedule_solution.schedule.shifts)  # Find max number of TAs per shift

        # Step 4: Open the CSV file for writing
        with open(filename, mode='w', newline='') as file:
            writer = csv.writer(file)

            # Step 5: Write the header row to the CSV
            writer.writerow(header)

            # Step 6: Write the rows for assigned TAs
            for row_index in range(max_rows):
                row = [row_index+1]  # The first column is for TA names and IDs
                for shift in shifts_sorted_by_start_time:
                    # Collect assigned TAs for this shift
                    assigned_tas = [f"{ta.name} - {ta.availability_as_dict.get(shift.series)} time" for ta in shift.assigned_tas]

                    # Prepare the cell content with TA name and ID, if available
                    if row_index < len(assigned_tas):
                        # Only show one TA per row in the corresponding shift column
                        cell_content = assigned_tas[row_index]
                    else:
                        cell_content = ""  # Empty cell if no TA is assigned to this row for the shift
                    row.append(cell_content)

                # Step 7: Write the row to the CSV
                writer.writerow(row)

        print(f"Report has been saved to {filename}")
