import logging
import os
import datetime
import csv
import openpyxl
import datetime as dt
import pandas as pd

from itertools import groupby
from copy import deepcopy
from typing import Dict, List, Any
from pathlib import Path
from datetime import timedelta, time, date
from argparse import Namespace


# Custom Imports 
from .domain import ShiftAssignment, Shift, TA, Timetable

# """Provides utilities for data extraction and reporting."""
class DataConstructor:
    """utilities for creating the timetable from the data folder."""
    def __init__(self, ta_csv_path, shift_csv_path, availability_folder, load=True):
        self.ta_csv_path: str = ta_csv_path
        self.shift_csv_path: str = shift_csv_path
        self.availability_folder: str = availability_folder
        # Data holders
        self.ta_dataframe: pd.DataFrame   = None  # Placeholder for TA data    (pandas frame)
        self.shift_dataframe: pd.DataFrame = None  # Placeholder for Shift data (pandas frame)
        self.timetable: Timetable  = None
        # intermediate data holders
        self.avialability_matrix: Dict[str, Dict[str, str]] = {}  # {macid: {shift_series: availability}}
        if load:
            self.create()

    def create(self):
        # Construction routine
        self.timetable = Timetable(id=0, tas=[], shifts=[], shift_assignments=[])
        self.load_data()
        self.create_ta_objects()
        self.create_shift_objects()
        self.create_shift_assignments()
        self.extract_availabilities()
        if (self.validate_ta_availability()):
            self.update_ta_availability()
            print(f"Data succussfully extracted :)")
        else: 
            print("\n========================================")
        return self.timetable
    
    def load_data(self):
        self.ta_dataframe    = pd.read_csv(self.ta_csv_path)
        self.shift_dataframe = pd.read_csv(self.shift_csv_path)

    def create_ta_objects(self):
        for index, row in self.ta_dataframe.iterrows():
            ta = TA(
                id=row['macid'],
                name=row['name'],
                required_shifts_per_week= row['req_shift_per_week'],
                is_grad_student= row['type'] == 'Grad',
                desired=[],
                undesired=[],
                unavailable=[]
            )
            self.timetable.tas.append(ta)

    def create_shift_objects(self):
        for index, row in self.shift_dataframe.iterrows():
            shift = Shift(
                id = str(index),
                alias = row['name'],
                series = row['series'],
                day_of_week = row['day_of_week'],
                shift_date = dt.datetime.strptime(row['date'], '%Y-%m-%d').date(),
                start_time = dt.datetime.strptime(row['start_time'], '%H:%M').time(),
                end_time = (dt.datetime.strptime(row['start_time'], '%H:%M') + timedelta(hours=row["duration"])).time() ,
                required_tas = row['req_ta_per_shift']
            )
            self.timetable.shifts.append(shift)

    def create_shift_assignments(self):
        for shift in self.timetable.shifts:
            for i in range(shift.required_tas):
                shift_assignment = ShiftAssignment(
                    id = f"{shift.series}_{i}",
                    shift = shift,
                    assigned_ta = None
                )
                self.timetable.shift_assignments.append(shift_assignment)  
  
    def extract_availabilities(self):
        print(self.availability_folder)
        filenames = self._find_xlsx_files(self.availability_folder)

        for file in filenames:
            mac_id, availability = self._extract_availability_from_xlsx(file)
            self.avialability_matrix[mac_id] = availability
        
        return self.avialability_matrix
    
    def _extract_availability_from_xlsx(self, file_path: str, availability_tbl_name: str = "availability_tbl"):
        # Load the workbook and select the sheet
        wb = openpyxl.load_workbook(file_path)
        ws = wb['availability']
        
        # Extracting TA details from the top of the sheet
        mac_id = ws['B1'].value
        ta_name = ws['B2'].value
        ta_id   = ws['B3'].value

        # # Print table names in the sheet
        # print("Tables in the sheet:", ws.tables.keys())

        # Check if availability_tbl_name exists
        if availability_tbl_name not in ws.tables:
            print("Table 'availability_tbl' not found.")

        tbl = ws.tables[availability_tbl_name]
        
        # Extract the table range
        table_range = tbl.ref  # e.g., "A5:C10"
        # print(f"Table '{availability_tbl_name}' range: {table_range}")

        # Extract the rows from the table range
        rows = ws[table_range]

        # Convert rows into a list of lists
        data = [[cell.value for cell in row] for row in rows]

        # Convert the data into a pandas DataFrame
        headers = data[0]       # First row contains headers
        table_df = pd.DataFrame(data[1:], columns=headers)  # Remaining rows are data
        
        availability = {}
        for index, row in table_df.iterrows():
            shift_series = row['Shift Series']
            shift_series_availability = row['Availability']
            availability[shift_series] = shift_series_availability   

        # print(availability)
        return mac_id, availability

    def validate_ta_availability(self, log_error=True):
        missing_ta = 0
        for ta in self.timetable.tas:
            if self.avialability_matrix.get(ta.id) is None:
                missing_ta += 1
                if log_error:
                    print("========================================")
                    print(f"Found no record for TA: {ta.name}")
                    print(f"MacID = {ta.id}")
                    print(f"missing {missing_ta}/{len(self.timetable.tas)} forms so far")
                    print("========================================")
            else:
                count_available = sum(1 for availability in self.avialability_matrix[ta.id] if availability != "Unavailable")
                if count_available < ta.required_shifts_per_week:
                    if log_error:
                        print(f"TA {ta.id} has insufficient availability.")
                        print(f"Available shifts: {count_available}")
                        print(f"Required shifts: {ta.required_shifts_per_week}")

                    return False
        return not missing_ta
    
    def update_ta_availability(self):
        
        if (self.validate_ta_availability(log_error=False)):
            for ta in self.timetable.tas:
                availability_as_dict = self.avialability_matrix.get(ta.id)

                if availability_as_dict is None:
                    print(f"TA {ta.id} {ta.name}has no availability data. Assumes all shifts are available.")
                    continue

                print(f"TA {ta.id} {ta.name} has availability data.")
                for shift_group, availability_status in availability_as_dict.items():
                    shift = next((shift for shift in self.timetable.shifts if shift.series == shift_group), None) # Find the shift group
                    if shift is None:
                        print(f"Shift group {shift_group} not found.")
                        continue

                    if availability_status == "Available":
                        ta.desired.append(shift)
                    elif availability_status == "Unavailable":
                        ta.unavailable.append(shift)
                    elif availability_status == "Undesired":
                        ta.undesired.append(shift)
                    else:
                        print(f"Unknown availability status: {availability_status}")      
            return True
        return False

    def _find_csv_files(self, directory):
        path = Path(directory)
        return [str(file) for file in path.glob("*.csv") if not file.name.startswith('~$')]  # Exclude files starting with ~$ and no recursive search

    def _find_xlsx_files(self, directory):
        path = Path(directory)
        return [str(file) for file in path.glob("*.xlsx") if not file.name.startswith('~$')]  # Exclude files starting with ~$ and no recursive search

# class Schedule_Utils:
# """Provides utilities for data extraction and reporting."""
#     def __init__(self, output_dir: str = "outputs"):
#         self.output_dir = output_dir

#     # Reporting Utils
#     def group_shifts_by_day_of_week(self, schedule_solution: Timetable):
#         workday_shifts = {}
#         for shift in schedule_solution.schedule.shifts:
#             if shift.day_of_week not in workday_shifts:
#                 workday_shifts[shift.day_of_week] = []
#             workday_shifts[shift.day_of_week].append(shift)

#         # Sort shifts for each workday based on start_time
#         for day in workday_shifts:
#             workday_shifts[day] = sorted(workday_shifts[day], key=lambda s: s.start_time)
#         return workday_shifts

#     def report(self, schedule_solution: Timetable, filename='ta_schedule_report.csv'):
#         # Step 1: Group shifts by day_of_week and sort them by start_time
#         filename = f"{self.output_dir}/{filename}"
#         workday_shifts = self.group_shifts_by_day_of_week(schedule_solution)

#         # Step 2: Prepare the header for the CSV file
#         days_of_week = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
#         header = ['Shift/TA'] + days_of_week  # First column will be shift names, rest are days

#         # Step 3: Determine the max number of rows across all workdays
#         max_rows = max(len(shifts) for shifts in workday_shifts.values())  # Find the max number of shifts per day

#         # Step 4: Open the CSV file for writing
#         with open(filename, mode='w', newline='') as file:
#             writer = csv.writer(file)

#             # Step 5: Write the header row to the CSV
#             writer.writerow(header)

#             # Step 6: Write the shift rows
#             for row_index in range(max_rows):
#                 row = ['']  # The first column is for shift names (e.g., Lab_0_0)
#                 for day in days_of_week:
#                     shifts = workday_shifts.get(day, [])
#                     if row_index < len(shifts):
#                         shift = shifts[row_index]
#                         # Collect assigned TAs for this shift
#                         assigned_tas = [f"{ta.id}_{ta.name}" for ta in shift.assigned_tas]
#                         # Prepare the cell content with shift name and assigned TAs
#                         cell_content = f"{shift.series}_{shift.id}\n" + "\n".join(assigned_tas) if assigned_tas else f"{shift.series}_{shift.id}\nNo TAs assigned"
#                     else:
#                         cell_content = ""  # Empty cell if no shift for this row
#                     row.append(cell_content)

#                 # Step 7: Write the row to the CSV
#                 writer.writerow(row)

#         print(f"Report has been saved to {filename}")

#     def report_timeseries(self, schedule_solution: Timetable, filename='ta_schedule_report_timeseries.csv'):
#         filename = f"{self.output_dir}/{filename}"
#         # Step 1: Extract and sort the shifts based on start time
#         shifts_sorted_by_start_time = sorted(schedule_solution.schedule.shifts, key=lambda s: s.start_time)

#         # Step 2: Prepare the header for the CSV file using the start time of each shift
#         header = ['TA Name/row']  # First column will be TA names and IDs
#         for shift in shifts_sorted_by_start_time:
#             # Format date and time as a string (assumes shift has a start_time field)
#             date_combined  = datetime.combine(shift.date, shift.start_time)
#             start_time_str = date_combined.strftime('%B %d %H:%M')
#             to_append = f"{shift.series} ({start_time_str})"
#             header.append(to_append)

#         # Step 3: Determine the max number of rows (based on the most TAs assigned to a shift)
#         max_rows = max(len(shift.assigned_tas) for shift in schedule_solution.schedule.shifts)  # Find max number of TAs per shift

#         # Step 4: Open the CSV file for writing
#         with open(filename, mode='w', newline='') as file:
#             writer = csv.writer(file)

#             # Step 5: Write the header row to the CSV
#             writer.writerow(header)

#             # Step 6: Write the rows for assigned TAs
#             for row_index in range(max_rows):
#                 row = [row_index+1]  # The first column is for TA names and IDs
#                 for shift in shifts_sorted_by_start_time:
#                     # Collect assigned TAs for this shift
#                     assigned_tas = [f"{ta.name} - {ta.availability_as_dict.get(shift.series)} time" for ta in shift.assigned_tas]

#                     # Prepare the cell content with TA name and ID, if available
#                     if row_index < len(assigned_tas):
#                         # Only show one TA per row in the corresponding shift column
#                         cell_content = assigned_tas[row_index]
#                     else:
#                         cell_content = ""  # Empty cell if no TA is assigned to this row for the shift
#                     row.append(cell_content)

#                 # Step 7: Write the row to the CSV
#                 writer.writerow(row)

#         print(f"Report has been saved to {filename}")


def id_generator():
    """Generates unique IDs for the shift assignments."""
    current = 0
    while True:
        yield str(current)
        current += 1

def initialize_logger(args: Namespace, logging_level = logging.INFO) -> logging.Logger:
    # Construct log directory path based on the constraint version argument
    variant: str = args.demo_data_select if not args.overwrite else f"overwrite"
    log_dir = os.path.join("logs", args.constraint_version, variant)
    os.makedirs(log_dir, exist_ok=True)

    # Use readable timestamp for the filename
    log_filename = f"timefold_assignment_matrix_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.log"
    log_path = os.path.join(log_dir, log_filename)

    logging.basicConfig(
        level=logging.INFO,
        # format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
        format="%(message)s",
        handlers=[
            logging.FileHandler(log_path, mode='w'),  # Save logs to a file, overwrite each run
            logging.StreamHandler()  # Also display logs in the console
        ]
    )
    # Create a logger instance
    LOGGER = logging.getLogger('app')

    LOGGER.info(f"...")
    LOGGER.info(f"Logger initialized")
    LOGGER.info(f"\t@ {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    LOGGER.info(f"\tConstraint version: {args.constraint_version}")
    LOGGER.info(f"\tData:               {variant}")
    LOGGER.info(f"...")

    return LOGGER

def get_handler_info(handler: logging.Handler) -> dict:
    """Returns a dictionary with information about the logging handler."""
    if not isinstance(handler, logging.Handler):
        raise ValueError("Handler must be an instance of logging.Handler")
    return {
        "type": type(handler).__name__,
        "level": logging.getLevelName(handler.level),
        "formatter": handler.formatter._fmt if handler.formatter else None,
        "output_file": getattr(handler, "baseFilename", None),  # Only FileHandlers have this
        "name": getattr(handler, "name", None)
    }

def get_handler_info_from_logger(logger: logging.Logger) -> List[Dict[str, Any]]:
    """Returns a list of dictionaries with information about the logging handlers of the given logger."""
    if not isinstance(logger, logging.Logger):
        raise ValueError("Logger must be an instance of logging.Logger")
    
    return [get_handler_info(handler) for handler in logger.handlers]

def create_logger_info(logger: logging.Logger | None) -> Dict[str, Any]:
    """Creates a dictionary with information about the logger."""
    if not isinstance(logger, logging.Logger):
        raise ValueError("Logger must be an instance of logging.Logger")
    
    return {
        "name": logger.name,
        "level": logging.getLevelName(logger.level),
        "effective_level": logging.getLevelName(logger.getEffectiveLevel()),
        "propagate": logger.propagate,
        "handlers": get_handler_info_from_logger(logger)
    }

def legacy_print_timetable(time_table: Timetable, logger: logging.Logger) -> None:

    LOGGER = logger
    LOGGER.info(f"Score: {time_table.score}")
    # LOGGER.info("=== Starting to print the timetable ===")

    
    tas = time_table.tas
    shift_groups = time_table.shifts
    shift_assignments = time_table.shift_assignments
    
    PADDDING = 10
    MIN_WIDTH = 22
    # Calculate the maximum width for the columns based on TA names
    column_width = max(max([len(ta.name) for ta in tas]) + PADDDING, MIN_WIDTH)  # Ensure a minimum width of 20 for better readability
    
    assignment_map = {
        (assignment.assigned_ta.name, assignment.shift.series, assignment.shift.start_time): assignment
        for assignment in shift_assignments
        if assignment.shift is not None and assignment.assigned_ta is not None
    }

    row_format = ("|{:<" + str(column_width) + "}") * (len(tas) + 1) + "|"
    sep_format = "+" + ((("-" * column_width) + "+") * (len(tas) + 1))

    LOGGER.info(sep_format)
    LOGGER.info(row_format.format('', *[f"{ta.name} (ID: {ta.id})" for ta in tas]))
    LOGGER.info(row_format.format('', *[f"requires" for ta in tas]))
    LOGGER.info(row_format.format('', *[f"  budget:      {ta.required_shifts_per_semester} " for ta in tas]))
    LOGGER.info(row_format.format('', *[f"  min_weekly: {ta.min_shifts_per_week} " for ta in tas]))
    LOGGER.info(row_format.format('', *[f"  max_weekly: {ta.max_shifts_per_week} " for ta in tas]))

    LOGGER.info(sep_format)

    ids = id_generator()
    for shift_group in shift_groups:
        def get_row_shifts():
            for ta in tas:
                yield assignment_map.get((ta.name, shift_group.series, shift_group.start_time),
                                     ShiftAssignment(id=next(ids), shift=shift_group, assigned_ta=None))

        # Logging the shift group
        row_shifts = [*get_row_shifts()]
        LOGGER.info(row_format.format(
            str(shift_group), 
            *[assignment.assigned_ta.name if assignment.assigned_ta is not None else " " 
              for assignment in row_shifts]
        ))

        LOGGER.info(row_format.format( 
            f"requires {shift_group.required_tas} TAs",
            *[assignment.assigned_ta.get_status_for_shift(shift_group) if assignment.assigned_ta is not None else " " 
              for assignment in row_shifts]
        ))

        LOGGER.info(sep_format)

    unassigned_shifts = [assignment for assignment in shift_assignments if assignment.shift is None or assignment.assigned_ta is None]
    if len(unassigned_shifts) > 0:
        LOGGER.info("")
        LOGGER.info("Unassigned shifts")
        for shift in unassigned_shifts:
            LOGGER.info(f'{shift.shift} - {shift.assigned_ta}')

    LOGGER.info("=== Finished printing the timetable ===")

def print_timetable(time_table: Timetable, logger: logging.Logger) -> None:
    LOGGER = logger
    LOGGER.info(f"Score: {time_table.score}")

    tas = time_table.tas
    # 1) Sort shifts by week_id
    shift_groups = sorted(time_table.shifts, key=lambda s: s.week_id)
    # 2) Build a lookup for assignments
    shift_assignments = time_table.shift_assignments
    assignment_map = {
        (asg.assigned_ta.name, asg.shift.series, asg.shift.start_time): asg
        for asg in shift_assignments
        if asg.assigned_ta is not None
    }

    # Table formatting
    PADDING = 10
    MIN_WIDTH = 22
    column_width = max(max(len(ta.name) for ta in tas) + PADDING, MIN_WIDTH)
    row_fmt = ("|{:<" + str(column_width) + "}") * (len(tas) + 1) + "|"
    sep_fmt = "+" + ((("-" * column_width) + "+") * (len(tas) + 1))

    # Header rows
    LOGGER.info(sep_fmt)
    LOGGER.info(row_fmt.format('', *[f"{ta.name} (ID: {ta.id})" for ta in tas]))
    LOGGER.info(row_fmt.format('', *["requires" for _ in tas]))
    LOGGER.info(row_fmt.format('', *[f"  budget: {ta.required_shifts_per_semester}" for ta in tas]))
    LOGGER.info(row_fmt.format('', *[f"  minWk:  {ta.min_shifts_per_week}"         for ta in tas]))
    LOGGER.info(row_fmt.format('', *[f"  maxWk:  {ta.max_shifts_per_week}"         for ta in tas]))
    LOGGER.info(sep_fmt)

    # 3) Group by week_id and print
    for week_id, shifts_in_week in groupby(shift_groups, key=lambda s: s.week_id):
        LOGGER.info(f"***** Week {week_id} *****")
        LOGGER.info(sep_fmt)

        for shift_group in shifts_in_week:
            # build the row of ShiftAssignment objects or empty placeholders
            row_asgs = [
                assignment_map.get((ta.name, shift_group.series, shift_group.start_time),
                                   ShiftAssignment(id="unassigned", shift=shift_group, assigned_ta=None))
                for ta in tas
            ]

            # 4a) Print the TA names or blank
            LOGGER.info(
                row_fmt.format(
                    str(shift_group),
                    *[asg.assigned_ta.name if asg.assigned_ta else "" for asg in row_asgs]
                )
            )
            # 4b) Print the status (Desired/Undesired/Unavailable/Neutral)
            LOGGER.info(
                row_fmt.format(
                    f"requires {shift_group.required_tas} TAs",
                    *[
                        asg.assigned_ta.get_status_for_shift(shift_group)
                        if asg.assigned_ta else ""
                        for asg in row_asgs
                    ]
                )
            )
            LOGGER.info(sep_fmt)

    # 5) Optionally list any completely unassigned slots
    unassigned = [
        asg for asg in shift_assignments
        if asg.assigned_ta is None
    ]
    if unassigned:
        LOGGER.info("\nUnassigned Shifts:")
        for asg in unassigned:
            LOGGER.info(f"  {asg.shift} → {asg.assigned_ta}")

def print_ta_availability(time_table: Timetable, logger: logging.Logger) -> None:
    """Prints the TA availability in a formatted way."""
    time_table_in = deepcopy(time_table)  # Ensure we do not modify the original timetable
    LOGGER = logger
    LOGGER.info("=== Starting to print the TA availability ===")
    
    tas:                    List[TA]                = time_table_in.tas
    shift_groups:           List[Shift]             = time_table_in.shifts
    shift_assignments_new:  List[ShiftAssignment]   = []                # a dummy list to hold the cross product of TAs and shifts
                                                                        # this is because we want to re-use the logic of the print_timetable function
    # Create a cross product of TAs and shifts
    ids = id_generator()
    for ta in tas:
        for shift in shift_groups:
            if (shift in ta.desired) or (shift in ta.undesired) or (shift in ta.unavailable): # do not add "neutral" shifts
                shift_assignments_new.append(ShiftAssignment(id=next(ids), shift=shift, assigned_ta=ta))

    time_table_in.shift_assignments = shift_assignments_new
    print_timetable(time_table_in, logger)

    LOGGER.info("=== Finished printing the TA availability ===")


# wrap the helper functions in a class
class HelperFunctions:
    """A class to wrap the helper functions for better organization."""
    
    @staticmethod
    def print_timetable(time_table: Timetable, logger: logging.Logger) -> None:
        """Prints the timetable in a formatted way."""
        print_timetable(time_table, logger)
    @staticmethod
    def initialize_logger(constraint_version: str = "default") -> logging.Logger:
        """Initializes the logger for the application."""
        return initialize_logger(constraint_version=constraint_version)
    
# example on how to use the HelperFunctions class
if __name__ == "__main__":
    logger  = HelperFunctions.initialize_logger()
    # Assuming `time_table` is an instance of `Timetable`
    logger_info=create_logger_info(logger)
    print(logger_info)