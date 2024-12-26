import logging
import os
import datetime
import csv
import openpyxl
import datetime as dt
import pandas as pd

from typing import Dict
from pathlib import Path
from datetime import timedelta, time, date


# Custom Imports 
from .domain import ShiftAssignment, Shift, TA, Timetable

class DataConstructor:
    """utilities for creating the timetable from the data folder."""
    def __init__(self, ta_csv_path, shift_csv_path, availability_folder, load=True):
        self.ta_csv_path: str = ta_csv_path
        self.shift_csv_path: str = shift_csv_path
        self.availability_folder: str = availability_folder
        # Data holders
        self.ta_dataframe: pd.DataFrame   = None  # Placeholder for TA data    (pandas frame)
        self.shift_dataframe: pd.DataFrame = None  # Placeholder for Shift data (pandas frame)
        self.timetable: Timetable  = None
        # intermediate data holders
        self.avialability_matrix: Dict[str, Dict[str, str]] = {}  # {macid: {shift_series: availability}}
        if load:
            self.create()

    def create(self):
        # Construction routine
        self.timetable = Timetable(id=0, tas=[], shifts=[], shift_assignments=[])
        self.load_data()
        self.create_ta_objects()
        self.create_shift_objects()
        self.create_shift_assignments()
        self.extract_availabilities()
        if (self.validate_ta_availability()):
            self.update_ta_availability()
            print(f"Data succussfully extracted :)")
        else: 
            print("\n========================================")
        return self.timetable
    
    def load_data(self):
        self.ta_dataframe    = pd.read_csv(self.ta_csv_path)
        self.shift_dataframe = pd.read_csv(self.shift_csv_path)

    def create_ta_objects(self):
        for index, row in self.ta_dataframe.iterrows():
            ta = TA(
                id=row['macid'],
                name=row['name'],
                required_shifts= row['req_shift_per_week'],
                is_grad_student= row['type'] == 'Grad',
                desired=[],
                undesired=[],
                unavailable=[]
            )
            self.timetable.tas.append(ta)

    def create_shift_objects(self):
        for index, row in self.shift_dataframe.iterrows():
            shift = Shift(
                id = str(index),
                alias = row['name'],
                series = row['series'],
                day_of_week = row['day_of_week'],
                shift_date = dt.datetime.strptime(row['date'], '%Y-%m-%d').date(),
                start_time = dt.datetime.strptime(row['start_time'], '%H:%M').time(),
                end_time = (dt.datetime.strptime(row['start_time'], '%H:%M') + timedelta(hours=row["duration"])).time() ,
                required_tas = row['req_ta_per_shift']
            )
            self.timetable.shifts.append(shift)

    def create_shift_assignments(self):
        for shift in self.timetable.shifts:
            for i in range(shift.required_tas):
                shift_assignment = ShiftAssignment(
                    id = f"{shift.series}_{i}",
                    shift = shift,
                    assigned_ta = None
                )
                self.timetable.shift_assignments.append(shift_assignment)  
  
    def extract_availabilities(self):
        print(self.availability_folder)
        filenames = self._find_xlsx_files(self.availability_folder)

        for file in filenames:
            mac_id, availability = self._extract_availability_from_xlsx(file)
            self.avialability_matrix[mac_id] = availability
        
        return self.avialability_matrix
    
    def _extract_availability_from_xlsx(self, file_path: str, availability_tbl_name: str = "availability_tbl"):
        # Load the workbook and select the sheet
        wb = openpyxl.load_workbook(file_path)
        ws = wb['availability']
        
        # Extracting TA details from the top of the sheet
        mac_id = ws['B1'].value
        ta_name = ws['B2'].value
        ta_id   = ws['B3'].value

        # # Print table names in the sheet
        # print("Tables in the sheet:", ws.tables.keys())

        # Check if availability_tbl_name exists
        if availability_tbl_name not in ws.tables:
            print("Table 'availability_tbl' not found.")

        tbl = ws.tables[availability_tbl_name]
        
        # Extract the table range
        table_range = tbl.ref  # e.g., "A5:C10"
        # print(f"Table '{availability_tbl_name}' range: {table_range}")

        # Extract the rows from the table range
        rows = ws[table_range]

        # Convert rows into a list of lists
        data = [[cell.value for cell in row] for row in rows]

        # Convert the data into a pandas DataFrame
        headers = data[0]       # First row contains headers
        table_df = pd.DataFrame(data[1:], columns=headers)  # Remaining rows are data
        
        availability = {}
        for index, row in table_df.iterrows():
            shift_series = row['Shift Series']
            shift_series_availability = row['Availability']
            availability[shift_series] = shift_series_availability   

        # print(availability)
        return mac_id, availability

    def validate_ta_availability(self, log_error=True):
        missing_ta = 0
        for ta in self.timetable.tas:
            if self.avialability_matrix.get(ta.id) is None:
                missing_ta += 1
                if log_error:
                    print("========================================")
                    print(f"Found no record for TA: {ta.name}")
                    print(f"MacID = {ta.id}")
                    print(f"missing {missing_ta}/{len(self.timetable.tas)} forms so far")
                    print("========================================")
            else:
                count_available = sum(1 for availability in self.avialability_matrix[ta.id] if availability != "Unavailable")
                if count_available < ta.required_shifts:
                    if log_error:
                        print(f"TA {ta.id} has insufficient availability.")
                        print(f"Available shifts: {count_available}")
                        print(f"Required shifts: {ta.required_shifts}")

                    return False
        return not missing_ta
    
    def update_ta_availability(self):
        
        if (self.validate_ta_availability(log_error=False)):
            for ta in self.timetable.tas:
                availability_as_dict = self.avialability_matrix.get(ta.id)

                if availability_as_dict is None:
                    print(f"TA {ta.id} {ta.name}has no availability data. Assumes all shifts are available.")
                    continue

                print(f"TA {ta.id} {ta.name} has availability data.")
                for shift_group, availability_status in availability_as_dict.items():
                    shift = next((shift for shift in self.timetable.shifts if shift.series == shift_group), None) # Find the shift group
                    if shift is None:
                        print(f"Shift group {shift_group} not found.")
                        continue

                    if availability_status == "Available":
                        ta.desired.append(shift)
                    elif availability_status == "Unavailable":
                        ta.unavailable.append(shift)
                    elif availability_status == "Undesired":
                        ta.undesired.append(shift)
                    else:
                        print(f"Unknown availability status: {availability_status}")      
            return True
        return False

    def _find_csv_files(self, directory):
        path = Path(directory)
        return [str(file) for file in path.glob("*.csv") if not file.name.startswith('~$')]  # Exclude files starting with ~$ and no recursive search

    def _find_xlsx_files(self, directory):
        path = Path(directory)
        return [str(file) for file in path.glob("*.xlsx") if not file.name.startswith('~$')]  # Exclude files starting with ~$ and no recursive search
