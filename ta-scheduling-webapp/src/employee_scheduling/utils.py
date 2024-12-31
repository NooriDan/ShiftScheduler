import logging
import os
import datetime
import csv
import openpyxl
import datetime as dt
import pandas as pd

from typing import Dict, List
from enum import Enum
from pathlib import Path
from datetime import timedelta, time, date, datetime
from dataclasses import dataclass, field

# Custom Imports 
from .domain import ShiftAssignment, Shift, TA, Timetable

DAY_START_TIME = time(14, 30)
DAY_END_TIME   = time(17, 30)

AFTERNOON_START_TIME =  time(18, 30)
AFTERNOON_END_TIME   =  time(21, 30)

def id_generator():
    """Generates unique IDs for the shift assignments."""
    current = 0
    while True:
        yield str(current)
        current += 1

def initialize_logger():
    # Configure the logging
    LOG_FILE = f"logs/timefold_webapp_{datetime.now().strftime('%y_%m_%d')}.log"

    # Create the logs directory if it doesn't exist
    if not os.path.exists("logs"):
        os.makedirs("logs")

    logging.basicConfig(
        level=logging.INFO,
        # format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
        format="%(asctime)s - %(name)s - %(message)s",
        handlers=[
            # logging.FileHandler(LOG_FILE, mode='w'),  # Save logs to a file, overwrite each run
            logging.StreamHandler()  # Also display logs in the console
        ]
    )
    return logging.getLogger('app')

class DataConstructor:
    """utilities for creating the timetable from the data folder."""
    def __init__(self, ta_csv_path, shift_csv_path, availability_folder, load=True):
        self.ta_csv_path: str = ta_csv_path
        self.shift_csv_path: str = shift_csv_path
        self.availability_folder: str = availability_folder
        # Data holders
        self.ta_dataframe: pd.DataFrame   = None  # Placeholder for TA data    (pandas frame)
        self.shift_dataframe: pd.DataFrame = None  # Placeholder for Shift data (pandas frame)
        self.timetable: Timetable  = None
        # intermediate data holders
        self.avialability_matrix: Dict[str, Dict[str, str]] = {}  # {macid: {shift_series: availability}}
        if load:
            self.create()

    def create(self):
        # Construction routine
        self.timetable = Timetable(id="DEFAULT", tas=[], shifts=[], shift_assignments=[])
        self.load_data()
        self.create_ta_objects()
        self.create_shift_objects()
        self.create_shift_assignments()
        self.extract_availabilities()
        if (self.validate_ta_availability()):
            self.update_ta_availability()
            print(f"Data succussfully extracted :)")
        else: 
            print("\n========================================")
        return self.timetable
    
    def load_data(self):
        self.ta_dataframe    = pd.read_csv(self.ta_csv_path)
        self.shift_dataframe = pd.read_csv(self.shift_csv_path)

    def create_ta_objects(self):
        for index, row in self.ta_dataframe.iterrows():
            ta = TA(
                id=row['macid'],
                name=row['name'],
                required_shifts= row['req_shift_per_week'],
                is_grad_student= row['type'] == 'Grad',
                desired=[],
                undesired=[],
                unavailable=[]
            )
            self.timetable.tas.append(ta)

    def create_shift_objects(self):
        for index, row in self.shift_dataframe.iterrows():
            shift = Shift(
                id = str(index),
                alias = row['name'],
                series = row['series'],
                day_of_week = row['day_of_week'],
                shift_date = dt.datetime.strptime(row['date'], '%Y-%m-%d').date(),
                start_time = dt.datetime.strptime(row['start_time'], '%H:%M').time(),
                end_time = (dt.datetime.strptime(row['start_time'], '%H:%M') + timedelta(hours=row["duration"])).time() ,
                required_tas = row['req_ta_per_shift']
            )
            self.timetable.shifts.append(shift)

    def create_shift_assignments(self):
        for shift in self.timetable.shifts:
            for i in range(shift.required_tas):
                shift_assignment = ShiftAssignment(
                    id = f"{shift.series}_{i}",
                    shift = shift,
                    assigned_ta = None
                )
                self.timetable.shift_assignments.append(shift_assignment)  
  
    def extract_availabilities(self):
        print(self.availability_folder)
        filenames = self._find_xlsx_files(self.availability_folder)

        for file in filenames:
            mac_id, availability = self._extract_availability_from_xlsx(file)
            self.avialability_matrix[mac_id] = availability
        
        return self.avialability_matrix
    
    def _extract_availability_from_xlsx(self, file_path: str, availability_tbl_name: str = "availability_tbl"):
        # Load the workbook and select the sheet
        wb = openpyxl.load_workbook(file_path)
        ws = wb['availability']
        
        # Extracting TA details from the top of the sheet
        mac_id = ws['B1'].value
        ta_name = ws['B2'].value
        ta_id   = ws['B3'].value

        # # Print table names in the sheet
        # print("Tables in the sheet:", ws.tables.keys())

        # Check if availability_tbl_name exists
        if availability_tbl_name not in ws.tables:
            print("Table 'availability_tbl' not found.")

        tbl = ws.tables[availability_tbl_name]
        
        # Extract the table range
        table_range = tbl.ref  # e.g., "A5:C10"
        # print(f"Table '{availability_tbl_name}' range: {table_range}")

        # Extract the rows from the table range
        rows = ws[table_range]

        # Convert rows into a list of lists
        data = [[cell.value for cell in row] for row in rows]

        # Convert the data into a pandas DataFrame
        headers = data[0]       # First row contains headers
        table_df = pd.DataFrame(data[1:], columns=headers)  # Remaining rows are data
        
        availability = {}
        for index, row in table_df.iterrows():
            shift_series = row['Shift Series']
            shift_series_availability = row['Availability']
            availability[shift_series] = shift_series_availability   

        # print(availability)
        return mac_id, availability

    def validate_ta_availability(self, log_error=True):
        missing_ta = 0
        for ta in self.timetable.tas:
            if self.avialability_matrix.get(ta.id) is None:
                missing_ta += 1
                if log_error:
                    print("========================================")
                    print(f"Found no record for TA: {ta.name}")
                    print(f"MacID = {ta.id}")
                    print(f"missing {missing_ta}/{len(self.timetable.tas)} forms so far")
                    print("========================================")
            else:
                count_available = sum(1 for availability in self.avialability_matrix[ta.id] if availability != "Unavailable")
                if count_available < ta.required_shifts:
                    if log_error:
                        print(f"TA {ta.id} has insufficient availability.")
                        print(f"Available shifts: {count_available}")
                        print(f"Required shifts: {ta.required_shifts}")

                    return False
        return not missing_ta
    
    def update_ta_availability(self):
        
        if (self.validate_ta_availability(log_error=False)):
            for ta in self.timetable.tas:
                availability_as_dict = self.avialability_matrix.get(ta.id)

                if availability_as_dict is None:
                    print(f"TA {ta.id} {ta.name}has no availability data. Assumes all shifts are available.")
                    continue

                print(f"TA {ta.id} {ta.name} has availability data.")
                for shift_group, availability_status in availability_as_dict.items():
                    shift = next((shift for shift in self.timetable.shifts if shift.series == shift_group), None) # Find the shift group
                    if shift is None:
                        print(f"Shift group {shift_group} not found.")
                        continue

                    if availability_status == "Available":
                        ta.desired.append(shift)
                    elif availability_status == "Unavailable":
                        ta.unavailable.append(shift)
                    elif availability_status == "Undesired":
                        ta.undesired.append(shift)
                    else:
                        print(f"Unknown availability status: {availability_status}")      
            return True
        return False

    def _find_csv_files(self, directory):
        path = Path(directory)
        return [str(file) for file in path.glob("*.csv") if not file.name.startswith('~$')]  # Exclude files starting with ~$ and no recursive search

    def _find_xlsx_files(self, directory):
        path = Path(directory)
        return [str(file) for file in path.glob("*.xlsx") if not file.name.startswith('~$')]  # Exclude files starting with ~$ and no recursive search

class DemoData(Enum):
    DemoA = "DemoA"
    DemoB = "DemoB"

def generate_demo_data(name: str = "DemoA") -> Timetable:
    if name == DemoData.DemoA:
        return _demo_data_A()
    if name == DemoData.DemoB:
        return _demo_data_B()
    else:
        raise ValueError(f"Unknown demo data name: {name}")

def _demo_data_A() -> Timetable:     
    ids = id_generator()
    shifts: List[Shift] = [
        Shift(id=next(ids), series="L07", day_of_week="Mon",start_time=DAY_START_TIME, end_time=DAY_END_TIME, required_tas=2, shiftDate=datetime.now().date()),
        Shift(id=next(ids), series="L08", day_of_week = "Mon", start_time=AFTERNOON_START_TIME, end_time=AFTERNOON_END_TIME, required_tas=3, shiftDate=datetime.now().date()),
        Shift(id=next(ids), series="L09", day_of_week="Tue", start_time=DAY_START_TIME, end_time=DAY_END_TIME, required_tas=2, shiftDate=datetime.now().date()+ timedelta(days=1)),
        Shift(id=next(ids), series="L10", day_of_week="Tue", start_time=AFTERNOON_START_TIME, end_time=AFTERNOON_END_TIME, required_tas=1, shiftDate=datetime.now().date()+ timedelta(days=1)),
        Shift(id=next(ids), series="L11", day_of_week="Thu", start_time=AFTERNOON_START_TIME, end_time=AFTERNOON_END_TIME, required_tas=2, shiftDate=datetime.now().date()+ timedelta(days=2)),
        ]
    
    ids = id_generator()
    course_tas: List[TA] = [
        TA(id = next(ids), 
            name = "M. Roghani", 
            required_shifts= 3, 
            unavailable= [shifts[4]], 
            desired    = [shifts[0], shifts[1], shifts[2]],
            undesired  = [shifts[3]]
        ),
        TA(id = next(ids), 
            name = "D. Noori", 
            required_shifts= 2, 
            unavailable= [shifts[0]], 
            desired    = [],
            undesired  = [shifts[2]]
        ),
        TA(id = next(ids), 
            name = "A. Gholami", 
            required_shifts= 1, 
            unavailable= [shifts[1]], 
            desired    = [shifts[0]],
            undesired  = [shifts[2]]
        ),
        TA(id = next(ids), 
            name = "M. Jafari", 
            required_shifts= 2, 
            unavailable= [shifts[3]], 
            desired    = [shifts[1]],
            undesired  = [shifts[2]]
        ),
        TA(id = next(ids), 
            name = "A. Athar", 
            required_shifts= 2, 
            unavailable= [], 
            desired    = [shifts[0]],
            undesired  = [shifts[2], shifts[1], shifts[3], shifts[4]]
        ),
    ]

    ids = id_generator()
    shift_assignments: List[ShiftAssignment] = []

    for shift in shifts:
        for i in range(shift.required_tas):
            shift_assignments.append(ShiftAssignment(id= next(ids), shift= shift, assigned_ta= None))

    return Timetable(
                    id= "DemoA", 
                    shifts=shifts, 
                    tas=course_tas, 
                    shift_assignments= shift_assignments
            )

def _demo_data_B() -> Timetable:
    ids = id_generator()
    shifts: List[Shift] = [
        Shift(id=next(ids),series="L01", day_of_week="Mon", start_time=DAY_START_TIME, end_time=DAY_END_TIME, required_tas=2),
        Shift(id=next(ids), series="L02", day_of_week="Mon", start_time=AFTERNOON_START_TIME, end_time=AFTERNOON_END_TIME, required_tas=3),
        Shift(id=next(ids), series="L03", day_of_week="Tue", start_time=DAY_START_TIME, end_time=DAY_END_TIME, required_tas=2),
        Shift(id=next(ids), series="L04", day_of_week="Tue", start_time=AFTERNOON_START_TIME, end_time=AFTERNOON_END_TIME, required_tas=1),
        Shift(id=next(ids), series="L05", day_of_week="Wed", start_time=DAY_START_TIME, end_time=DAY_END_TIME, required_tas=2),
        Shift(id=next(ids), series="L06", day_of_week="Wed", start_time=AFTERNOON_START_TIME, end_time=AFTERNOON_END_TIME, required_tas=2),
        Shift(id=next(ids), series="L07", day_of_week="Thu", start_time=AFTERNOON_START_TIME, end_time=AFTERNOON_END_TIME, required_tas=3),
        Shift(id=next(ids), series="L08", day_of_week="Thu", start_time=AFTERNOON_START_TIME, end_time=AFTERNOON_END_TIME, required_tas=2),
        Shift(id=next(ids), series="L09", day_of_week="Fri", start_time=DAY_START_TIME, end_time=DAY_END_TIME, required_tas=2),
        Shift(id=next(ids), series="L10", day_of_week="Fri", start_time=AFTERNOON_START_TIME, end_time=AFTERNOON_END_TIME, required_tas=2),
        ]
    # total number of shifts = 21

    ids = id_generator()
    course_tas: List[TA] = [
        TA(id = next(ids), 
            name = "M. Roghani", 
            required_shifts= 3, 
            unavailable= [shifts[6], shifts[7], shifts[8]], 
            desired    = [shifts[0], shifts[1], shifts[2]],
            undesired  = [shifts[3], shifts[4], shifts[5]]
        ),

        TA(id = next(ids), 
            name = "D. Noori", 
            required_shifts= 2, 
            unavailable= shifts[0:5], 
            desired    = [shifts[6], shifts[7]],
            undesired  = [shifts[8]]
        ),

        TA(id = next(ids), 
            name = "A. Gholami", 
            required_shifts= 3, 
            unavailable= [shifts[1], shifts[3], shifts[5]], 
            desired    = [shifts[0], shifts[8]],
            undesired  = [shifts[4], shifts[6], shifts[7], shifts[9]]
        ),

        TA(id = next(ids), 
            name = "M. Jafari", 
            required_shifts= 2, 
            unavailable= [shifts[3]], 
            desired    = [shifts[1]],
            undesired  = [shifts[2]]
        ),

        TA(id = next(ids), 
            name = "A. Athar", 
            required_shifts= 2, 
            unavailable= [], 
            desired    = [shifts[0]],
            undesired  = [shifts[2], shifts[1], shifts[3], shifts[4]]
        ),

        TA(id = next(ids),
            name = "A. Anderson",
            required_shifts= 3,
            unavailable= shifts[6:-1],
            desired=shifts[0:3],
            undesired=[]
        ),

        TA(id = next(ids),
            name = "B. Brown",
            required_shifts= 3,
            unavailable= [],
            desired=[shifts[0]],
            undesired=[shifts[2], shifts[1], shifts[3], shifts[4]]
        ),

        TA(id = next(ids),
            name = "C. Campbell",
            required_shifts= 1,
            unavailable= [shifts[3], shifts[4]],
            desired=[shifts[0], shifts[1]],
            undesired=[shifts[2], shifts[1], shifts[5]]
        ),

        TA(id = next(ids),
            name = "D. Davis",
            required_shifts= 2,
            unavailable= [],
            desired=[shifts[0]],
            undesired=[shifts[2], shifts[1], shifts[3], shifts[4]]
        ),
    ]
    # total number of TAs = 5, total number of required shifts = 21

    ids = id_generator()
    shift_assignments: List[ShiftAssignment] = []

    for shift in shifts:
        for i in range(shift.required_tas):
            shift_assignments.append(ShiftAssignment(id= next(ids), shift= shift, assigned_ta= None))

    return Timetable(
                id= "Demo_B", 
                shifts=shifts, 
                tas=course_tas, 
                shift_assignments= shift_assignments
        )

